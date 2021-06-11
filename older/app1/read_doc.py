# -*- coding:utf-8 -*-
from openpyxl import *
from pprint import pprint

def read_doc(filename, sheetname, mode=None, max_line=None):
    cells = []
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
    y = sheet.max_column
    cell = []
    for j in range(1, y + 1):
        cell.append(sheet.cell(1, j).value)
    cells.append(cell)

    if mode == None:
        print("Mode None")
    else:
        for i in range(x - max_line + 1, x + 1):
            cell = []
            for j in range(1, y + 1):
                cell.append(sheet.cell(i, j).value)
            cells.append(cell)
    return cells


if __name__ == "__main__":
    x = read_doc(filename='doc/release.xlsx', sheetname='plan', mode=1, max_line=8)
    pprint(x)
