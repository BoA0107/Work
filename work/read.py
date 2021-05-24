# -*- coding:utf-8 -*-

from openpyxl import *


def read_doc(filename, sheetname, max_line=0):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
    y = sheet.max_column

    content = []
    if max_line == 0:
        for i in range(1, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)

    else:
        c = []
        for j in range(1, y + 1):
            c.append(sheet.cell(1, j).value)
        content.append(c)
        for i in range(x + 1 - max_line, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)
    return content

    # plan_info = read_doc(filename=r"doc/release.xlsx", sheetname="plan", max_line=8)
    #
    # print(plan_info)


# make dct
def make_dct(lst):
    dct = {}
    for i in range(len(lst)):
        x, y = lst[i]
        if (y == None) and (x != None):
            key = x
            dct[key] = {}
            for j in range(i + 1, len(lst)):
                x, y = lst[j]
                if y != None:
                    dct[key][x] = y
                else:
                    break
    return dct
