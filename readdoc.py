# -*- coding:utf-8 -*-

from openpyxl import *


def read_doc(filename, sheetname, max_line=0):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
    y = sheet.max_column


def make_dct(filename, sheetname):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
