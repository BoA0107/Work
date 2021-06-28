# -*- coding: UTF-8 -*-

from openpyxl import load_workbook


# read excel
def read_doc(filename, sheetname, max_line=0):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    x = sheet.max_row
    y = sheet.max_column

    content = []
    if max_line == 0:
        for i in range(1, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)
    else:
        c = []
        for j in range(1, y + 1):
            c.append(sheet.cell(1, j).value)
        content.append(c)
        for i in range(x + 1 - max_line, x + 1):
            c = []
            for j in range(1, y + 1):
                c.append(sheet.cell(i, j).value)
            content.append(c)

    return content


# make dct
def make_dct(lst):
    dct = {}
    for i in range(len(lst)):
        x, y = lst[i]
        if (y == None) and (x != None):
            key = x
            dct[key] = {}
            for j in range(i + 1, len(lst)):
                x, y = lst[j]
                if y != None:
                    dct[key][x] = y
                else:
                    break
    return dct


# read text
def read_txt(filename):
    f = []
    with open(filename) as file:
        x = file.readlines()
        for i in range(len(x)):
            f.append(x[i].strip('\n'))
    return f


#read single excel

def read_singel(filename,sheetname):
    wb = load_workbook(filename)
    sheet = wb.get_sheet_by_name(sheetname)
    content=[]
    x = sheet.max_row
    for i in range(1, x + 1):
        content.append(sheet.cell(i, 1).value)
    return content